import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from typing import List


EXCEL_HEADERS = [
    "Country", "ProductType", "ProductGrade", "ProductPriceMin", "ProductPriceMax",
    "LengthUnit", "WeightUnit",
    "LengthMin", "LengthMax",
    "WidthMin", "WidthMax", "HeightMin", "HeightMax",
    "WeightMin", "WeightMax", "Price", "Currency", "Remark"
]


def export_rules_to_excel(rules: List[dict]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipping Rules"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, rule in enumerate(rules, 2):
        values = [
            rule.get("country", ""),
            rule.get("productType", ""),
            rule.get("productGrade", ""),
            rule.get("productPriceMin", 0),
            rule.get("productPriceMax", 0),
            rule.get("lengthUnit", "cm"),
            rule.get("weightUnit", "g"),
            rule.get("lengthMin", 0),
            rule.get("lengthMax", 0),
            rule.get("widthMin", 0),
            rule.get("widthMax", 0),
            rule.get("heightMin", 0),
            rule.get("heightMax", 0),
            rule.get("weightMin", 0),
            rule.get("weightMax", 0),
            rule.get("price", 0),
            rule.get("currency", ""),
            rule.get("remark", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 4, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel_to_rules(file_bytes: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active

    rules = []
    header_row = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(header_row, row))
        if not row_data.get("Country") and not row_data.get("country"):
            continue

        try:
            rule = {
                "country": str(row_data.get("Country", row_data.get("country", "")) or ""),
                "productType": str(row_data.get("ProductType", row_data.get("productType", "")) or ""),
                "productGrade": str(row_data.get("ProductGrade", row_data.get("productGrade", "")) or ""),
                "productPriceMin": float(row_data.get("ProductPriceMin", row_data.get("productPriceMin", 0)) or 0),
                "productPriceMax": float(row_data.get("ProductPriceMax", row_data.get("productPriceMax", 0)) or 0),
                "lengthUnit": str(row_data.get("LengthUnit", row_data.get("lengthUnit", "cm")) or "cm"),
                "weightUnit": str(row_data.get("WeightUnit", row_data.get("weightUnit", "g")) or "g"),
                "lengthMin": float(row_data.get("LengthMin", row_data.get("lengthMin", 0)) or 0),
                "lengthMax": float(row_data.get("LengthMax", row_data.get("lengthMax", 0)) or 0),
                "widthMin": float(row_data.get("WidthMin", row_data.get("widthMin", 0)) or 0),
                "widthMax": float(row_data.get("WidthMax", row_data.get("widthMax", 0)) or 0),
                "heightMin": float(row_data.get("HeightMin", row_data.get("heightMin", 0)) or 0),
                "heightMax": float(row_data.get("HeightMax", row_data.get("heightMax", 0)) or 0),
                "weightMin": float(row_data.get("WeightMin", row_data.get("weightMin", 0)) or 0),
                "weightMax": float(row_data.get("WeightMax", row_data.get("weightMax", 0)) or 0),
                "price": float(row_data.get("Price", row_data.get("price", 0)) or 0),
                "currency": str(row_data.get("Currency", row_data.get("currency", "USD")) or "USD"),
                "remark": str(row_data.get("Remark", row_data.get("remark", "")) or ""),
            }
            rules.append(rule)
        except (ValueError, TypeError):
            continue

    return rules
